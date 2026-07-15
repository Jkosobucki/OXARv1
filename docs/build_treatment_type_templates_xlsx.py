#!/usr/bin/env python3
"""OXar_Treatment_Type_Templates.xlsx + treatment-type-templates.ui.json

One template-definition tab per treatment type, now carrying the full UI binding for each field:
Entity · Attribute (logical name) · Data type · Control · Relationship / Option set · Required ·
Sample value · Src. This is directly consumable by a UI generator / form engine. Also emits an
enriched JSON with the same metadata. Mirrors the TEMPLATES/META in
reference/oxar_prototype_screens.html. Independent of OX.gp.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEEP="2F7D76"; SOFT="EAF5F3"; INK="221F20"; WHITE="FFFFFF"; ZEBRA="F5F8F7"; AMBER="C9821C"; MUTED="6C7C7A"
wb=openpyxl.Workbook()
thin=Side(style="thin",color="D9E4E2"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hf=Font(bold=True,color=WHITE,size=10); hfill=PatternFill("solid",fgColor=DEEP)
title=Font(bold=True,color=INK,size=15); sub=Font(italic=True,color=MUTED,size=10); lab=Font(bold=True,color=DEEP,size=10)
wrap=Alignment(vertical="top",wrap_text=True)

PATIENT="Priya Sample (DEMO) · NHR-0042 · F · DOB 12 Mar 1991"
PARTNER="Alex Sample (DEMO) · NHR-0043 · M · DOB 04 Aug 1988"
ALL_TABS=["Planning","Folliculogram","Drugs","Oocyte Retrieval","Semen","Lab Results",
          "Embryo Transfer","IUI/DI","Gametes","Stored Sample","Outcome Data","Checklists"]
TAB_ENTITY={  # which entity each cycle tab primarily binds to (for the UI)
 "Planning":"bcrm_treatment_cycle","Folliculogram":"bcrm_fertilitytreatment","Drugs":"AppointmentDosage",
 "Oocyte Retrieval":"bcrm_opu","Semen":"bcrm_semenprep","Lab Results":"bcrm_eggdetail",
 "Embryo Transfer":"bcrm_embryotransfer","IUI/DI":"bcrm_insemination","Gametes":"bcrm_egg",
 "Stored Sample":"bcrm_egg / bcrm_cryolocation","Outcome Data":"bcrm_pregnancy_outcome_details",
 "Checklists":"bcrm_consentchecklist"}

# ---- Field binding metadata: label -> (entity, attribute, dataType, control, relationship/optionset, src) ----
# src: code=confirmed in OxArBackendReact; form=from the cycle form (confirm logical name); new=Embryo Module entity
M = {
 "Treatment Name":("bcrm_treatment_cycle","bcrm_name","Text","Textbox","—","code"),
 "Treatment Type":("bcrm_treatment_cycle","bcrm_treatmenttype","Choice","Dropdown","OptionSet: bcrm_treatmenttype","code"),
 "Cycle Sequence No":("bcrm_treatment_cycle","bcrm_cyclesequencenumber","AutoNumber","Read-only","—","form"),
 "Billing Option":("bcrm_treatment_cycle","bcrm_treatmentbillingoption","Choice","Dropdown","OptionSet","form"),
 "ANZARD Cycle Type":("bcrm_treatment_cycle","bcrm_anzardcycletype","Choice","Dropdown","OptionSet","form"),
 "Treatment Template":("bcrm_treatment_cycle","bcrm_treatmenttemplate","Lookup","Lookup","N:1 -> bcrm_eventtemplate","new"),
 "Patient":("bcrm_treatment_cycle","bcrm_patient","Lookup","Lookup","N:1 -> contact","code"),
 "Partner":("bcrm_treatment_cycle","bcrm_partner","Lookup","Lookup","N:1 -> contact","code"),
 "Cycle Doctor":("bcrm_treatment_cycle","bcrm_cycledoctor","Lookup","Lookup","N:1 -> systemuser/bcrm_staff","form"),
 "Cycle Nurse":("bcrm_treatment_cycle","bcrm_cyclenurse","Lookup","Lookup","N:1 -> systemuser/bcrm_staff","form"),
 "Patient RX Group":("bcrm_treatment_cycle","bcrm_patientrxgroup","Choice","Dropdown","OptionSet","form"),
 "Embryologist":("bcrm_thawevent","bcrm_embryologist","Lookup","Lookup","N:1 -> systemuser/bcrm_staff","new"),
 "Witness":("bcrm_thawevent","bcrm_witness","Lookup","Lookup","N:1 -> systemuser/bcrm_staff (independent)","new"),
 "Managing Clinic":("bcrm_treatment_cycle","bcrm_managingclinic","Lookup","Lookup","N:1 -> account/bcrm_clinic","form"),
 "Procedure Clinic":("bcrm_treatment_cycle","bcrm_procedureclinic","Lookup","Lookup","N:1 -> account/bcrm_clinic","form"),
 "Expected Start":("bcrm_treatment_cycle","bcrm_expectedstartdate","DateTime","Date picker","—","form"),
 "First Injection":("bcrm_fertilitytreatment","bcrm_firstinjectiondate","DateTime","Date picker","1:1 stim detail","code"),
 "Expected OPU":("bcrm_treatment_cycle","bcrm_expectedtreatmentdate","DateTime","Date picker","—","form"),
 "Lining Check":("bcrm_fertilitytreatment","bcrm_liningcheckdate","DateTime","Date picker","1:1 stim detail","new"),
 "Expected Transfer":("bcrm_treatment_cycle","bcrm_expectedtreatmentdate","DateTime","Date picker","—","form"),
 "Expected Thaw":("bcrm_thawevent","bcrm_scheduledthawdate","DateTime","Date picker","1:N per cycle","new"),
 "LH Surge":("bcrm_fertilitytreatment","bcrm_lhsurgedate","DateTime","Date picker","1:1 stim detail","new"),
 "Expected Insemination":("bcrm_insemination","bcrm_date","DateTime","Date picker","1:N per cycle","new"),
 "Baseline Scan":("bcrm_fertilitytreatment","bcrm_dateofbaselinescan","DateTime","Date picker","1:1 stim detail","code"),
 "Day-10 Scan":("bcrm_fertilitytreatment","bcrm_day10scandate","DateTime","Date picker","1:1 stim detail","new"),
 "Period Started":("bcrm_treatment_cycle","bcrm_expecteddateperiodstarted","DateTime","Date picker","—","form"),
 "Actual Period Date":("bcrm_treatment_cycle","bcrm_actualperioddate","DateTime","Date picker","—","form"),
 "Female Diagnosis":("bcrm_treatment_cycle","bcrm_treatment_details_diagnosis","Choice","Dropdown","OptionSet / Lookup","code"),
 "Male Diagnosis":("bcrm_treatment_cycle","bcrm_treatment_details_sec_diagnosis","Choice","Dropdown","OptionSet / Lookup","code"),
 "Comment":("bcrm_treatment_cycle","bcrm_treatment_details_other_diagnosis","Multiline Text","Textarea","—","code"),
 "Reason":("bcrm_fertilitytreatment","bcrm_treatmentfeatures","Multiline Text","Textarea","—","code"),
 "AMH Validated":("bcrm_fertilitytreatment","bcrm_amh_validated","Two Options","Toggle","—","code"),
 "Event Template":("bcrm_treatment_cycle","bcrm_eventtemplate","Lookup","Lookup","N:1 -> bcrm_eventtemplate (ApplyToCycle seeds events)","new"),
 "Research Project":("bcrm_treatment_cycle","bcrm_researchproject","Lookup","Lookup","N:1 -> bcrm_researchproject","new"),
 "Sperm Source":("bcrm_fertilitytreatment","bcrm_spermsource","Choice","Dropdown","OptionSet","code"),
 "Insemination Method":("bcrm_insemination","bcrm_method","Choice","Dropdown","OptionSet (IVF/ICSI)","new"),
 "No. Embryos to Replace":("bcrm_fertilitytreatment","bcrm_no_of_embryos_to_be_replaced","Whole Number","Number","—","code"),
 "Freeze Method":("bcrm_freezeevent","bcrm_freezemethod","Choice","Dropdown","OptionSet","new"),
 "Target eggs to store":("bcrm_treatment_cycle","bcrm_targeteggstostore","Text","Textbox","—","new"),
 "Endometrial Prep":("bcrm_fertilitytreatment","bcrm_endometrialprep","Choice","Dropdown","OptionSet (Medicated/Natural)","new"),
 "Selected Embryo":("bcrm_embryotransfer","bcrm_embryo","Lookup","Lookup","N:1 -> bcrm_egg (embryo)","new"),
 "Frozen Oocytes to Thaw":("bcrm_thawevent","bcrm_materialrecord","Lookup","Lookup","N:1 -> bcrm_egg (oocyte)","new"),
 "Prepared Motile Count":("bcrm_semenprep","bcrm_count","Decimal","Number","—","new"),
 "Trigger":("bcrm_fertilitytreatment","bcrm_timeoftrigger","DateTime","Date picker","—","code"),
 "Drug":("AppointmentDosage","bcrm_drug","Lookup / Text","Lookup","1:N drug rows","code"),
 "Monitoring":("bcrm_fertilitytreatment","bcrm_folliculogram_comments","Multiline Text","Textarea","—","code"),
 "Hand-off":("bcrm_treatment_cycle","bcrm_handoffnote","Multiline Text","Textarea","—","new"),
 "Specimen":("bcrm_egg","bcrm_eggs_id","Lookup","Lookup","N:1 -> bcrm_egg","new"),
 "Refreeze Method":("bcrm_freezeevent","bcrm_freezemethod","Choice","Dropdown","OptionSet","new"),
 "Storage Location":("bcrm_cryolocation","bcrm_cryolocationid","Lookup","Lookup","N:1 -> bcrm_cryolocation","new"),
}
def meta(label):
    return M.get(label, ("bcrm_treatment_cycle","(confirm)","Text","Textbox","—","new"))

def fld(l,v,r=""): return (l,v,r)
TEMPLATES={
"IVF":{"purpose":"Full stimulated cycle — collect eggs, fertilise, transfer / freeze.","partner":True,"note":"",
 "steps":["Consent & 3PC","Stimulation","Folliculogram","Trigger","OPU","Semen Prep","Insemination (ICSI)","Fertilisation","Culture D2-7","Transfer","Freeze surplus","Outcome"],
 "hide":["IUI/DI"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","IVF — Antagonist protocol","Yes"),fld("Treatment Type","IVF","Yes"),fld("Cycle Sequence No","CY-1183"),fld("Billing Option","Self-funded","Yes"),fld("ANZARD Cycle Type","Stimulated — fresh"),fld("Treatment Template","Standard ICSI, Day 5")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Partner",PARTNER),fld("Cycle Doctor","Dr M. Osei","Yes"),fld("Cycle Nurse","R. Nolan"),fld("Patient RX Group","Group B","Yes")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("Procedure Clinic","OX.ar City Lab"),fld("Expected Start","14 Jul 2026"),fld("First Injection","16 Jul 2026"),fld("Expected OPU","28 Jul 2026")]),
  ("Infertility Diagnosis",[fld("Female Diagnosis","Tubal factor"),fld("Male Diagnosis","Mild oligospermia"),fld("Comment","—")]),
  ("Plan",[fld("Event Template","IVF Standard"),fld("Research Project","PROTO time-lapse study"),fld("Sperm Source","Partner (fresh)"),fld("Insemination Method","ICSI"),fld("No. Embryos to Replace","1")]),
 ]},
"Egg Only":{"purpose":"Egg freezing / fertility preservation — collect & vitrify oocytes (no fertilisation).","partner":False,"note":"",
 "steps":["Consent & 3PC","Stimulation","Folliculogram","Trigger","OPU","Denuding / MII","Vitrify Oocytes","Storage"],
 "hide":["Embryo Transfer","IUI/DI","Lab Results","Outcome Data"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","Egg Freezing","Yes"),fld("Treatment Type","Egg Only","Yes"),fld("Cycle Sequence No","CY-1184"),fld("Billing Option","Self-funded","Yes"),fld("ANZARD Cycle Type","Stimulated — freeze all"),fld("Treatment Template","Egg Freezing")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Cycle Doctor","Dr M. Osei","Yes"),fld("Cycle Nurse","R. Nolan"),fld("Patient RX Group","Group B","Yes")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("Procedure Clinic","OX.ar City Lab"),fld("Expected Start","14 Jul 2026"),fld("First Injection","16 Jul 2026"),fld("Expected OPU","28 Jul 2026")]),
  ("Indication",[fld("Reason","Elective fertility preservation"),fld("AMH Validated","Yes — 18.2 pmol/L")]),
  ("Plan",[fld("Event Template","Egg Freezing"),fld("Research Project","—"),fld("Freeze Method","Vitrification"),fld("Target eggs to store","≥ 10 MII")]),
 ]},
"Embryo Transfer":{"purpose":"Frozen embryo transfer (FET) — thaw a stored embryo and transfer.","partner":True,"note":"",
 "steps":["Consent & 3PC","Endometrial Prep / Lining","Select Embryo","Embryo Thaw","3PC","Transfer","Outcome"],
 "hide":["Oocyte Retrieval","IUI/DI","Lab Results"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","FET — Medicated","Yes"),fld("Treatment Type","Embryo Transfer","Yes"),fld("Cycle Sequence No","CY-1185"),fld("Billing Option","Medicare + gap","Yes"),fld("ANZARD Cycle Type","FET"),fld("Treatment Template","FET (Frozen Embryo Transfer)")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Partner",PARTNER),fld("Cycle Doctor","Dr M. Osei","Yes"),fld("Cycle Nurse","R. Nolan"),fld("Patient RX Group","Group B","Yes")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("Lining Check","22 Jul 2026"),fld("Expected Transfer","29 Jul 2026")]),
  ("Plan",[fld("Event Template","FET"),fld("Endometrial Prep","Medicated (HRT)"),fld("Selected Embryo","EMB-1042 · Blast 4AA · Day 5"),fld("No. Embryos to Replace","1")]),
 ]},
"VOT":{"purpose":"Vitrified Oocyte Thaw — thaw frozen eggs → ICSI → culture → transfer.","partner":True,"note":"VOT is a site-specific acronym — confirm exact meaning with MIVF/SME.",
 "steps":["Consent & 3PC","Select Frozen Eggs","Oocyte Thaw","Survival Check","Semen Prep","ICSI","Fertilisation","Culture","Transfer / Freeze","Outcome"],
 "hide":["Oocyte Retrieval","Folliculogram","IUI/DI"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","VOT — thaw & ICSI","Yes"),fld("Treatment Type","VOT","Yes"),fld("Cycle Sequence No","CY-1186"),fld("Billing Option","Self-funded","Yes"),fld("ANZARD Cycle Type","Thaw oocytes"),fld("Treatment Template","Vitrified Oocyte Thaw")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Partner",PARTNER),fld("Cycle Doctor","Dr M. Osei","Yes"),fld("Patient RX Group","Group B","Yes")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("Expected Thaw","26 Jul 2026"),fld("Expected Transfer","31 Jul 2026")]),
  ("Plan",[fld("Event Template","VOT"),fld("Frozen Oocytes to Thaw","6 (MII)"),fld("Sperm Source","Partner (fresh)"),fld("Insemination Method","ICSI"),fld("No. Embryos to Replace","1")]),
 ]},
"IUI":{"purpose":"Intrauterine insemination — prepared sperm placed in the uterus (no egg collection).","partner":True,"note":"",
 "steps":["Consent & 3PC","OI / Natural","Folliculogram","Trigger","Semen Prep","IUI Insemination","Outcome"],
 "hide":["Oocyte Retrieval","Embryo Transfer","Lab Results","Stored Sample"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","IUI — Natural cycle","Yes"),fld("Treatment Type","IUI","Yes"),fld("Cycle Sequence No","CY-1187"),fld("Billing Option","Self-funded","Yes"),fld("ANZARD Cycle Type","IUI"),fld("Treatment Template","IUI")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Partner",PARTNER),fld("Cycle Doctor","Dr M. Osei","Yes"),fld("Cycle Nurse","R. Nolan"),fld("Patient RX Group","Group A","Yes")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("LH Surge","24 Jul 2026"),fld("Expected Insemination","25 Jul 2026")]),
  ("Infertility Diagnosis",[fld("Female Diagnosis","Unexplained"),fld("Male Diagnosis","Normal parameters")]),
  ("Plan",[fld("Event Template","IUI"),fld("Sperm Source","Partner (prepared)"),fld("Prepared Motile Count","18 M/mL"),fld("Trigger","hCG")]),
 ]},
"Ovulation Induction":{"purpose":"Drug-induced ovulation + timed intercourse (monitoring only, no lab procedure).","partner":False,"note":"",
 "steps":["Consent","OI Drugs","Folliculogram","Trigger","Timed Intercourse","Outcome"],
 "hide":["Oocyte Retrieval","Embryo Transfer","IUI/DI","Gametes","Lab Results","Stored Sample"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","Ovulation Induction — Letrozole","Yes"),fld("Treatment Type","Ovulation Induction","Yes"),fld("Cycle Sequence No","CY-1188"),fld("Billing Option","Self-funded","Yes"),fld("Treatment Template","Ovulation Induction")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Cycle Doctor","Dr M. Osei","Yes"),fld("Cycle Nurse","R. Nolan")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("Baseline Scan","14 Jul 2026"),fld("Day-10 Scan","24 Jul 2026")]),
  ("Infertility Diagnosis",[fld("Female Diagnosis","PCOS — anovulation")]),
  ("Plan",[fld("Drug","Letrozole 5 mg (days 3-7)"),fld("Monitoring","Serial follicle scans"),fld("Trigger","hCG when lead follicle ≥ 18 mm")]),
 ]},
"Tracking":{"purpose":"Cycle monitoring / scan tracking only (no procedure).","partner":False,"note":"",
 "steps":["Folliculogram Tracking","LH & Trigger Monitoring","Outcome / Hand-off"],
 "hide":["Oocyte Retrieval","Embryo Transfer","IUI/DI","Gametes","Lab Results","Stored Sample","Checklists"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","Cycle Tracking","Yes"),fld("Treatment Type","Tracking","Yes"),fld("Cycle Sequence No","CY-1189"),fld("Treatment Template","Cycle Tracking")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Cycle Doctor","Dr M. Osei"),fld("Cycle Nurse","R. Nolan")]),
  ("Clinic & Dates",[fld("Managing Clinic","OX.ar City (DEMO)"),fld("Period Started","10 Jul 2026"),fld("Actual Period Date","10 Jul 2026")]),
  ("Plan",[fld("Monitoring","Serial scans + LH"),fld("Hand-off","to IUI / IVF if indicated")]),
 ]},
"Thaw and Refreeze":{"purpose":"Retrieve a stored specimen → thaw → (± biopsy) → refreeze → storage.","partner":True,"note":"",
 "steps":["Consent & 3PC","Select Specimen","Thaw","Assess (± Biopsy)","Refreeze","Storage","Witnessing"],
 "hide":["Folliculogram","Drugs","Oocyte Retrieval","Embryo Transfer","IUI/DI","Lab Results","Outcome Data"],
 "sections":[
  ("Cycle Information",[fld("Treatment Name","Thaw & Refreeze — PGT","Yes"),fld("Treatment Type","Thaw and Refreeze","Yes"),fld("Cycle Sequence No","CY-1190"),fld("Treatment Template","Thaw and Refreeze")]),
  ("People",[fld("Patient",PATIENT,"Yes"),fld("Partner",PARTNER),fld("Embryologist","A. Kaur","Yes"),fld("Witness","R. Nolan","Yes")]),
  ("Specimen & Storage",[fld("Specimen","EMB-1043 · Blast 3BB"),fld("Reason","Biopsy for PGT-A"),fld("Refreeze Method","Vitrification"),fld("Storage Location","TANK-01 · C1 · G2 · S5")]),
 ]},
}

COLS=["Section","Field","Entity","Attribute (logical)","Data type","Control","Relationship / Option set","Required","Sample value","Src"]
WID=[18,22,26,32,14,13,34,9,34,6]
def style_header(ws,row,cols,widths):
    for j,c in enumerate(cols,1):
        cell=ws.cell(row,j,c); cell.font=hf; cell.fill=hfill; cell.border=border; cell.alignment=wrap
        ws.column_dimensions[get_column_letter(j)].width=widths[j-1]

# Index
ix=wb.active; ix.title="Index"; ix.sheet_view.showGridLines=False
ix["A1"]="OX.ar — Treatment-Type Templates (UI binding)"; ix["A1"].font=Font(bold=True,size=16,color=INK)
ix["A2"]=("One tab per treatment type = the template that drives the Treatment Planning UI. Each field carries "
          "its Entity, Attribute (logical name), Data type, Control and Relationship so it can be bound directly. "
          "Machine-readable copy: treatment-type-templates.ui.json. Synthetic test couple. Independent of OX.gp.")
ix["A2"].font=sub; ix.merge_cells("A2:F2"); ix["A2"].alignment=wrap; ix.row_dimensions[2].height=56
style_header(ix,4,["#","Treatment Type","Purpose","Uses partner?","Note"],[5,24,52,14,40])
for i,(t,tpl) in enumerate(TEMPLATES.items()):
    r=5+i
    for j,v in enumerate([str(i+1),t,tpl["purpose"],"Yes" if tpl["partner"] else "No",tpl["note"] or "—"],1):
        c=ix.cell(r,j,v); c.alignment=wrap; c.border=border
        if r%2==0: c.fill=PatternFill("solid",fgColor=ZEBRA)
ix.freeze_panes="A5"
tr=5+len(TEMPLATES)+2
ix.cell(tr,1,"Test couple (synthetic):").font=Font(bold=True)
ix.cell(tr+1,1,"Patient — "+PATIENT); ix.cell(tr+2,1,"Partner — "+PARTNER)
ix.cell(tr+4,1,"Src legend:").font=Font(bold=True)
ix.cell(tr+5,1,"code = attribute confirmed in OxArBackendReact  ·  form = from the cycle form (confirm logical name)  ·  new = new Embryo-Module entity/field")

# JSON accumulator
J={"version":"1.0","cycleEntity":"bcrm_treatment_cycle","typeField":"bcrm_treatmenttype",
   "testCouple":{"patient":PATIENT,"partner":PARTNER},
   "tabEntities":TAB_ENTITY,"allTabs":ALL_TABS,"templates":{}}

for t,tpl in TEMPLATES.items():
    ws=wb.create_sheet(t[:31]); ws.sheet_view.showGridLines=False
    ws["A1"]=t+" — Template Definition (UI binding)"; ws["A1"].font=title
    ws["A2"]=tpl["purpose"]; ws["A2"].font=sub; ws.merge_cells("A2:E2"); ws["A2"].alignment=wrap
    ws["A3"]="Uses partner: "+("Yes" if tpl["partner"] else "No")+("   ·   ⚠ "+tpl["note"] if tpl["note"] else "")
    ws["A3"].font=Font(size=10,color=(AMBER if tpl["note"] else MUTED)); ws.merge_cells("A3:E3"); ws["A3"].alignment=wrap
    ws["A5"]="Plan steps (seed the worklist):"; ws["A5"].font=lab
    ws["A6"]="  →  ".join(tpl["steps"]); ws.merge_cells("A6:J6"); ws["A6"].alignment=wrap; ws.row_dimensions[6].height=28
    hr=8; style_header(ws,hr,COLS,WID); r=hr+1
    jsecs=[]
    for sec,fields in tpl["sections"]:
        jf=[]
        for k,(fl,val,req) in enumerate(fields):
            e,a,dt,ctrl,rel,src=meta(fl)
            row=[sec if k==0 else "",fl,e,a,dt,ctrl,rel,req or "",val,src]
            for j,v in enumerate(row,1):
                cell=ws.cell(r,j,v); cell.border=border; cell.alignment=wrap
                if r%2==0: cell.fill=PatternFill("solid",fgColor=ZEBRA)
                if j==1 and k==0: cell.font=Font(bold=True,color=DEEP)
                if j in (3,4): cell.font=Font(name="Consolas",size=10)
                if j==10 and v=="new": cell.font=Font(color=DEEP,bold=True,size=9)
                if j==10 and v=="form": cell.font=Font(color=AMBER,bold=True,size=9)
            r+=1
            jf.append({"label":fl,"entity":e,"attribute":a,"dataType":dt,"control":ctrl,
                       "relationship":rel,"required":(req=="Yes"),"visible":True,"sampleValue":val,"src":src})
        jsecs.append({"name":sec,"fields":jf})
    # tabs table with entity binding
    r+=1
    for j,h in enumerate(["Cycle tab","Show / Hide","Binds to entity"],1):
        c=ws.cell(r,j,h); c.font=hf; c.fill=hfill; c.border=border
    r+=1
    for tab in ALL_TABS:
        shown=tab not in tpl["hide"]
        ws.cell(r,1,tab).border=border
        c=ws.cell(r,2,"SHOW" if shown else "HIDE"); c.border=border
        c.font=Font(bold=True,color=(DEEP if shown else "CF4A4A"),size=10)
        c.fill=PatternFill("solid",fgColor=(SOFT if shown else "FBE9E9"))
        ce=ws.cell(r,3,TAB_ENTITY.get(tab,"")); ce.border=border; ce.font=Font(name="Consolas",size=9)
        r+=1
    ws.freeze_panes="A9"
    J["templates"][t]={"purpose":tpl["purpose"],"usesPartner":tpl["partner"],"note":tpl["note"],
        "steps":tpl["steps"],"tabs":{"show":[x for x in ALL_TABS if x not in tpl["hide"]],"hide":tpl["hide"]},
        "sections":jsecs}

wb.save("/workspace/oxarv1/docs/OXar_Treatment_Type_Templates.xlsx")
with open("/workspace/oxarv1/docs/templates/treatment-type-templates.ui.json","w") as fp:
    json.dump(J,fp,indent=2,ensure_ascii=False)
print("saved xlsx sheets:", wb.sheetnames)
print("saved json templates:", list(J["templates"].keys()))
