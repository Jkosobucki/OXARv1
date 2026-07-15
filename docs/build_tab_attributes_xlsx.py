#!/usr/bin/env python3
"""Generate the OX.ar Treatment Cycle — Tabs & Attributes workbook.

One sheet per Treatment Cycle tab, listing the attributes (fields) with source entity,
relationship, data type and confirmation status. Grounded in the OxArBackendReact code
(bcrm_treatment_cycle / bcrm_fertilitytreatment / bcrm_pregnancy_outcome_details) and the
Embryo Module data model (01-DATA-MODEL.md). Independent of OX.gp.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AQUA = "45A59D"; DEEP = "2F7D76"; SOFT = "EAF5F3"; INK = "221F20"
WHITE = "FFFFFF"; ZEBRA = "F5F8F7"; AMBER = "C9821C"

wb = openpyxl.Workbook()

thin = Side(style="thin", color="D9E4E2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DEEP)
title_font = Font(name="Calibri", bold=True, color=INK, size=14)
sub_font = Font(name="Calibri", italic=True, color="6C7C7A", size=10)
wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")

COLS = ["Section", "Field (label)", "Logical name", "Data type", "Source entity",
        "Relationship to cycle", "Required", "Src", "Notes"]
WIDTHS = [20, 30, 34, 15, 26, 30, 9, 6, 46]

# Src legend: code = confirmed in backend code; form = from the Treatment Cycle form label
# (logical name to confirm); new = new entity/field from the Embryo Module data model.

def sheet(name, subtitle, rows):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = name; ws["A1"].font = title_font
    ws["A2"] = subtitle; ws["A2"].font = sub_font
    hr = 4
    for j, c in enumerate(COLS, 1):
        cell = ws.cell(hr, j, c); cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = wrap; cell.border = border
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS[j-1]
    for i, r in enumerate(rows):
        rr = hr + 1 + i
        for j, val in enumerate(r, 1):
            cell = ws.cell(rr, j, val); cell.alignment = wrap; cell.border = border
            if rr % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
            if j == 8 and val == "new":
                cell.font = Font(color=DEEP, bold=True, size=10)
            if j == 8 and val == "form":
                cell.font = Font(color=AMBER, bold=True, size=10)
    ws.freeze_panes = ws.cell(hr + 1, 1)
    ws.row_dimensions[1].height = 20
    return ws

# ---------------- Overview / index sheet ----------------
ov = wb.active; ov.title = "Tabs (read me)"; ov.sheet_view.showGridLines = False
ov["A1"] = "OX.ar — Treatment Cycle tabs & attributes"; ov["A1"].font = Font(bold=True, size=16, color=INK)
ov["A2"] = ("Redesigned tab order: Treatment Planning (consolidates today's Overview cycle info) "
            "then Three Point Check, then the clinical/lab tabs. Show/hide per treatment type is in "
            "06-TREATMENT-TYPE-TEMPLATES.md. Independent of OX.gp.")
ov["A2"].font = sub_font; ov.merge_cells("A2:F2"); ov["A2"].alignment = wrap
ov.row_dimensions[2].height = 42
idx_cols = ["#", "Tab", "Purpose", "Primary entity", "Key related entities"]
idx_w = [5, 26, 40, 26, 40]
for j, c in enumerate(idx_cols, 1):
    cell = ov.cell(4, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
    ov.column_dimensions[get_column_letter(j)].width = idx_w[j-1]
TABS = [
    ("1", "Treatment Planning", "Cycle set-up: who/what/when + plan (template, research). Consolidates today's Overview.", "bcrm_treatment_cycle", "contact, clinic, systemuser, bcrm_eventtemplate, bcrm_researchproject"),
    ("2", "Three Point Check", "Identity + consent verification gate before any critical step.", "bcrm_threepointcheck (new)", "bcrm_treatment_cycle, bcrm_embryoevent, systemuser"),
    ("3", "Folliculogram", "Stimulation monitoring: scans, ovaries, LH, trigger.", "bcrm_fertilitytreatment", "bcrm_treatment_cycle"),
    ("4", "Drugs / Stimulation", "Drug regimen & doses.", "AppointmentDosage", "bcrm_treatment_cycle, bcrm_fertilitytreatment"),
    ("5", "Oocyte Retrieval (OPU)", "Egg collection + MII/MI/GV maturity split.", "bcrm_opu (new)", "bcrm_treatment_cycle, bcrm_egg, systemuser"),
    ("6", "Semen / Andrology", "Semen prep: collection/thaw, count/motility/morphology.", "bcrm_semenprep (new)", "bcrm_treatment_cycle, bcrm_egg"),
    ("7", "Insemination & Fertilisation", "IVF/ICSI + PN/PB fertilisation check.", "bcrm_insemination (new)", "bcrm_eggdetail, bcrm_egg, bcrm_semenprep"),
    ("8", "Embryo Development", "Day 0-7 grade / fragmentation / stage / fate.", "bcrm_eggdetail", "bcrm_egg (embryo)"),
    ("9", "Embryo Transfer", "Fresh/FET transfer procedure + outcome.", "bcrm_embryotransfer (new)", "bcrm_egg, systemuser, bcrm_treatment_cycle"),
    ("10", "IUI / DI", "Intrauterine insemination procedure.", "bcrm_insemination (new)", "bcrm_semenprep, bcrm_treatment_cycle"),
    ("11", "Gametes", "Specimen master view (egg/sperm/embryo).", "bcrm_egg", "contact, bcrm_treatment_cycle, bcrm_cryolocation"),
    ("12", "Cryo Storage — Tanks & Inventory", "Tank graphical view + position hierarchy + inventory.", "bcrm_cryotank / bcrm_cryolocation (new)", "bcrm_egg, bcrm_site"),
    ("13", "Freeze / Thaw", "Cryopreservation & warming events.", "bcrm_freezeevent / bcrm_thawevent (new)", "bcrm_egg, bcrm_cryolocation, systemuser"),
    ("14", "Outcome Data", "Pregnancy test + delivery outcome.", "bcrm_pregnancy_outcome_details", "bcrm_treatment_cycle, contact"),
    ("15", "Checklists / Consent", "Consent/screening/approval prerequisites.", "bcrm_consentchecklist (new)", "bcrm_treatment_cycle, systemuser"),
    ("16", "Recommendation", "ET advice / scratch / clinical recommendations.", "bcrm_fertilitytreatment", "bcrm_treatment_cycle"),
    ("17", "General Information", "Referring centre, complications, features.", "bcrm_treatment_cycle / bcrm_fertilitytreatment", "-"),
]
for i, r in enumerate(TABS):
    rr = 5 + i
    for j, v in enumerate(r, 1):
        cell = ov.cell(rr, j, v); cell.alignment = wrap; cell.border = border
        if rr % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ZEBRA)
ov.freeze_panes = "A5"

# ---------------- Per-tab attribute sheets ----------------
# Row = (Section, Field label, Logical name, Data type, Source entity, Relationship, Required, Src, Notes)

sheet("1. Treatment Planning",
      "Consolidates today's Overview cycle info + a Plan section. Primary entity: bcrm_treatment_cycle.",
      [
        ("Cycle Information", "Treatment Name", "bcrm_name", "Text", "bcrm_treatment_cycle", "field on cycle", "Yes", "code", "Primary name / auto-number"),
        ("Cycle Information", "Treatment Type", "bcrm_treatmenttype", "Choice", "bcrm_treatment_cycle", "field on cycle", "Yes", "code", "Drives the display template (06-TREATMENT-TYPE-TEMPLATES.md)"),
        ("Cycle Information", "Cycle Sequence Number", "bcrm_cyclesequencenumber", "Whole Number / AutoNumber", "bcrm_treatment_cycle", "field on cycle", "No", "form", "System-generated; read-only"),
        ("Cycle Information", "Treatment Billing Option", "bcrm_treatmentbillingoption", "Choice", "bcrm_treatment_cycle", "field on cycle", "Yes", "form", "Confirm logical name"),
        ("Cycle Information", "Treatment Template", "bcrm_treatmenttemplate", "Lookup", "bcrm_eventtemplate", "N:1 -> bcrm_eventtemplate", "No", "new", "Selects the event/display template to apply"),
        ("Cycle Information", "Anzard Cycle Type", "bcrm_anzardcycletype", "Choice", "bcrm_treatment_cycle", "field on cycle", "Yes", "form", "ANZARD reporting classification"),
        ("Cycle Information", "Cycle Instructions", "bcrm_cycleinstructions", "Multiline Text", "bcrm_treatment_cycle", "field on cycle", "No", "form", ""),
        ("People", "Patient", "bcrm_patient", "Lookup", "contact", "N:1 -> contact", "Yes", "code", ""),
        ("People", "Partner", "bcrm_partner", "Lookup", "contact", "N:1 -> contact", "No", "code", "Show when partner gametes are used"),
        ("People", "Patient Group", "bcrm_patientgroup", "Choice", "bcrm_treatment_cycle", "field on cycle", "No", "form", ""),
        ("People", "Patient RX Group", "bcrm_patientrxgroup", "Choice", "bcrm_treatment_cycle", "field on cycle", "Yes", "form", ""),
        ("People", "Cycle Doctor", "bcrm_cycledoctor", "Lookup", "systemuser / bcrm_staff", "N:1 -> staff", "Yes", "form", ""),
        ("People", "Cycle Nurse", "bcrm_cyclenurse", "Lookup", "systemuser / bcrm_staff", "N:1 -> staff", "No", "form", ""),
        ("Clinic", "Managing Clinic", "bcrm_managingclinic", "Lookup", "account / bcrm_clinic", "N:1 -> clinic", "No", "form", ""),
        ("Clinic", "Procedure Clinic", "bcrm_procedureclinic", "Lookup", "account / bcrm_clinic", "N:1 -> clinic", "No", "form", ""),
        ("Dates", "Expected Start Date", "bcrm_expectedstartdate", "DateTime", "bcrm_treatment_cycle", "field on cycle", "No", "form", ""),
        ("Dates", "Expected Treatment Date", "bcrm_expectedtreatmentdate", "DateTime", "bcrm_treatment_cycle", "field on cycle", "No", "form", ""),
        ("Dates", "Treatment Date", "bcrm_treatmentdate", "DateTime", "bcrm_treatment_cycle", "field on cycle (header)", "No", "form", ""),
        ("Dates", "Expected Date Period Started", "bcrm_expecteddateperiodstarted", "DateTime", "bcrm_treatment_cycle", "field on cycle (header)", "No", "form", ""),
        ("Dates", "Actual Period Date", "bcrm_actualperioddate", "DateTime", "bcrm_treatment_cycle", "field on cycle (header)", "No", "form", ""),
        ("Dates", "First Injection Date", "bcrm_firstinjectiondate", "DateTime", "bcrm_fertilitytreatment", "1:1 stim details", "No", "code", ""),
        ("Dates", "Date of Baseline Scan", "bcrm_dateofbaselinescan", "DateTime", "bcrm_fertilitytreatment", "1:1 stim details", "No", "code", ""),
        ("Diagnosis", "Female Infertility Diagnosis 1", "bcrm_treatment_details_diagnosis", "Choice / Lookup", "bcrm_treatment_cycle", "field on cycle", "No", "code", ""),
        ("Diagnosis", "Male Infertility Diagnosis 1", "bcrm_treatment_details_sec_diagnosis", "Choice / Lookup", "bcrm_treatment_cycle", "field on cycle", "No", "code", ""),
        ("Diagnosis", "Infertility Diagnosis Comment", "bcrm_treatment_details_other_diagnosis", "Multiline Text", "bcrm_treatment_cycle", "field on cycle", "No", "code", ""),
        ("Plan", "Event Template (apply)", "bcrm_eventtemplate", "Lookup", "bcrm_eventtemplate", "N:1; ApplyToCycle seeds events", "No", "new", "Seeds bcrm_embryoevent worklist for the cycle"),
        ("Plan", "Research Project", "bcrm_researchproject", "Lookup", "bcrm_researchproject", "N:1 -> research project", "No", "new", "Drives research banner (M18-NEW)"),
        ("Plan", "Research Enrolled", "bcrm_researchenrolled", "Two Options", "bcrm_treatment_cycle", "field on cycle", "No", "new", ""),
        ("Plan", "Sperm Source", "bcrm_spermsource", "Choice / Text", "bcrm_fertilitytreatment", "1:1 stim details", "No", "code", "Partner / Donor / Surgical"),
        ("Plan", "Type of Cycle", "bcrm_type_of_cycle", "Choice", "bcrm_fertilitytreatment", "1:1 stim details", "No", "code", ""),
        ("Plan", "Age at Treatment", "bcrm_ageattreatment", "Whole Number", "bcrm_fertilitytreatment", "1:1 stim details", "No", "code", "Derived from patient DOB"),
        ("Plan", "Deferred Reason", "bcrm_deferredreason", "Choice / Lookup", "bcrm_treatment_cycle", "field on cycle", "No", "form", "Show only when cycle deferred"),
      ])

sheet("2. Three Point Check",
      "Identity + consent verification gate before any critical step. New entity bcrm_threepointcheck (01-DATA-MODEL.md 3.3).",
      [
        ("Check", "Check ID", "bcrm_threepointcheckid", "AutoNumber", "bcrm_threepointcheck", "primary key", "Yes", "new", "3PC-{seq}"),
        ("Check", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1 -> cycle", "Yes", "new", ""),
        ("Check", "Related Event", "bcrm_relatedevent", "Lookup", "bcrm_embryoevent", "N:1 -> event being gated", "No", "new", "The critical step this check unlocks"),
        ("Verification", "Patient Verified", "bcrm_patientverified", "Two Options", "bcrm_threepointcheck", "field", "Yes", "new", "Identifier 1"),
        ("Verification", "Partner Verified", "bcrm_partnerverified", "Two Options", "bcrm_threepointcheck", "field", "No", "new", "Identifier 2 (when applicable)"),
        ("Verification", "Consent Verified", "bcrm_consentverified", "Two Options", "bcrm_threepointcheck", "field", "Yes", "new", "Identifier 3"),
        ("People", "Verified By", "bcrm_verifiedby", "Lookup", "systemuser / bcrm_staff", "N:1 -> staff (performer)", "Yes", "new", ""),
        ("People", "Witness", "bcrm_witness", "Lookup", "systemuser / bcrm_staff", "N:1 -> staff (independent)", "Yes", "new", "MUST differ from Verified By (enforced server-side)"),
        ("Check", "Verification Date", "bcrm_verificationdate", "DateTime", "bcrm_threepointcheck", "field", "Yes", "new", ""),
        ("Check", "Status", "bcrm_status", "Choice", "bcrm_threepointcheck", "field", "Yes", "new", "Pending / Passed / Failed - blocks procedure until Passed"),
        ("Check", "Notes", "bcrm_notes", "Multiline Text", "bcrm_threepointcheck", "field", "No", "new", ""),
      ])

sheet("3. Folliculogram",
      "Stimulation monitoring. Primary entity: bcrm_fertilitytreatment (1:1 with the cycle).",
      [
        ("Monitoring", "Folliculogram data", "bcrm_folliculogram", "Multiline Text / JSON", "bcrm_fertilitytreatment", "field", "No", "code", "FolliculagramJSON in DTO"),
        ("Monitoring", "Left Ovary", "bcrm_leftovary", "Text", "bcrm_fertilitytreatment", "field", "No", "code", "Follicle measurements"),
        ("Monitoring", "Right Ovary", "bcrm_rightovary", "Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Monitoring", "LH", "bcrm_lh", "Text / Decimal", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Monitoring", "Uterus", "bcrm_uterus", "Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Monitoring", "First Ext OS-Fundus Length", "bcrm_first_eentityt_os_fundus_length", "Text", "bcrm_fertilitytreatment", "field", "No", "code", "Inherited spelling - keep on the wire"),
        ("Stimulation", "Drug Given", "bcrm_drug_given", "Text / Lookup", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Stimulation", "Dose", "bcrm_dose_v1", "Text / Decimal", "bcrm_fertilitytreatment", "field", "No", "code", "bcrm_dose also present"),
        ("Stimulation", "AMH Validated", "bcrm_amh_validated", "Two Options / Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Trigger", "Time of Trigger", "bcrm_timeoftrigger", "DateTime", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Location", "Stimulation Location", "bcrm_stimulation_location", "Text / Lookup", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Location", "Egg Collection Location", "bcrm_egg_collection_location", "Text / Lookup", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Location", "Referring Centre", "bcrm_referring_centre", "Text / Lookup", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Notes", "Folliculogram Comments", "bcrm_folliculogram_comments", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Notes", "Issues", "bcrm_issues", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Notes", "Sperm Source (Folliculogram)", "bcrm_sperm_source_folliculogram", "Choice", "bcrm_fertilitytreatment", "field", "No", "code", ""),
      ])

sheet("4. Drugs - Stimulation",
      "Drug regimen & doses. Primary entity: AppointmentDosage (drug schedule) + stim fields on bcrm_fertilitytreatment.",
      [
        ("Regimen", "Drug", "bcrm_drug", "Lookup / Text", "AppointmentDosage", "1:N drug rows per cycle", "Yes", "code", "AppointmentDosage controller exists"),
        ("Regimen", "Dose", "bcrm_dose", "Text / Decimal", "AppointmentDosage", "field", "Yes", "code", ""),
        ("Regimen", "Route", "bcrm_route", "Choice", "AppointmentDosage", "field", "No", "form", ""),
        ("Regimen", "Frequency", "bcrm_frequency", "Choice / Text", "AppointmentDosage", "field", "No", "form", ""),
        ("Regimen", "Start Date", "bcrm_startdate", "DateTime", "AppointmentDosage", "field", "No", "form", ""),
        ("Regimen", "End Date", "bcrm_enddate", "DateTime", "AppointmentDosage", "field", "No", "form", ""),
        ("Summary", "Drug Given (summary)", "bcrm_drug_given", "Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Summary", "Dose (summary)", "bcrm_dose_v1", "Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
      ])

sheet("5. Oocyte Retrieval (OPU)",
      "Egg collection. New entity bcrm_opu (01-DATA-MODEL.md 3.4). Produces oocyte bcrm_egg rows.",
      [
        ("OPU", "OPU ID", "bcrm_opuid", "AutoNumber", "bcrm_opu", "primary key", "Yes", "new", "OPU-{seq}"),
        ("OPU", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1 -> cycle", "Yes", "new", ""),
        ("OPU", "Procedure Date", "bcrm_proceduredate", "DateTime", "bcrm_opu", "field", "Yes", "new", ""),
        ("People", "Doctor", "bcrm_doctor", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("People", "Embryologist", "bcrm_embryologist", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("People", "Witness", "bcrm_witness", "Lookup", "systemuser / bcrm_staff", "N:1 (independent)", "Yes", "new", ""),
        ("Counts", "Follicles Aspirated", "bcrm_folliclesaspirated", "Whole Number", "bcrm_opu", "field", "No", "new", ""),
        ("Counts", "Eggs Retrieved", "bcrm_eggsretrieved", "Whole Number", "bcrm_opu", "field", "Yes", "new", "= Number of Eggs Collected on cycle"),
        ("Maturity", "MII", "bcrm_mii", "Whole Number", "bcrm_opu", "field", "No", "new", "Mature - ICSI eligible"),
        ("Maturity", "MI", "bcrm_mi", "Whole Number", "bcrm_opu", "field", "No", "new", ""),
        ("Maturity", "GV", "bcrm_gv", "Whole Number", "bcrm_opu", "field", "No", "new", ""),
        ("Notes", "Complications", "bcrm_complications", "Multiline Text", "bcrm_opu", "field", "No", "new", ""),
        ("Cycle roll-up", "Number of Eggs Collected", "bcrm_numberofeggscollected", "Whole Number", "bcrm_fertilitytreatment", "roll-up onto cycle", "No", "code", ""),
      ])

sheet("6. Semen - Andrology",
      "Semen preparation. New entity bcrm_semenprep (01-DATA-MODEL.md 3.5).",
      [
        ("Prep", "Semen Prep ID", "bcrm_semenprepid", "AutoNumber", "bcrm_semenprep", "primary key", "Yes", "new", "SP-{seq}"),
        ("Prep", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1", "Yes", "new", ""),
        ("Source", "Semen Source", "bcrm_semensource", "Choice", "bcrm_semenprep", "field", "Yes", "new", "Partner / Donor / Surgical (TESE/PESA)"),
        ("Source", "Fresh / Frozen", "bcrm_freshfrozen", "Choice", "bcrm_semenprep", "field", "Yes", "new", ""),
        ("Source", "Frozen Sample Record", "bcrm_frozensamplerecord", "Lookup", "bcrm_egg", "N:1 -> stored straw", "No", "new", "Marks owner straw thawed"),
        ("Dates", "Collection Date", "bcrm_collectiondate", "DateTime", "bcrm_semenprep", "field", "No", "new", ""),
        ("Dates", "Thaw Date", "bcrm_thawdate", "DateTime", "bcrm_semenprep", "field", "No", "new", ""),
        ("Analysis", "Volume (ml)", "bcrm_volume", "Decimal", "bcrm_semenprep", "field", "No", "new", ""),
        ("Analysis", "Count (x10^6/ml)", "bcrm_count", "Decimal", "bcrm_semenprep", "field", "No", "new", ""),
        ("Analysis", "Motility (%)", "bcrm_motility", "Whole Number", "bcrm_semenprep", "field", "No", "new", ""),
        ("Analysis", "Morphology (%)", "bcrm_morphology", "Whole Number", "bcrm_semenprep", "field", "No", "new", ""),
        ("Use", "Quantity Used", "bcrm_quantityused", "Decimal", "bcrm_semenprep", "field", "No", "new", ""),
        ("Use", "Quality Grade", "bcrm_qualitygrade", "Choice / Text", "bcrm_semenprep", "field", "No", "new", ""),
        ("Use", "Used For", "bcrm_usedfor", "Choice", "bcrm_semenprep", "field", "No", "new", "IVF / ICSI / IUI"),
      ])

sheet("7. Insemination - Fertilisation",
      "Insemination event + PN/PB fertilisation check. bcrm_insemination (new) + bcrm_eggdetail assessments.",
      [
        ("Insemination", "Insemination ID", "bcrm_inseminationid", "AutoNumber", "bcrm_insemination", "primary key", "Yes", "new", "INS-{seq}"),
        ("Insemination", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1", "Yes", "new", ""),
        ("Insemination", "Method", "bcrm_method", "Choice", "bcrm_insemination", "field", "Yes", "new", "IVF / ICSI"),
        ("Insemination", "Date", "bcrm_date", "DateTime", "bcrm_insemination", "field", "Yes", "new", ""),
        ("Insemination", "Embryologist", "bcrm_embryologist", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("Insemination", "Semen Prep", "bcrm_semenprep", "Lookup", "bcrm_semenprep", "N:1", "No", "new", ""),
        ("Insemination", "No of Eggs / Successful", "bcrm_noofeggs / bcrm_successfuleggs", "Whole Number", "bcrm_insemination", "field", "No", "new", ""),
        ("Fertilisation", "PN Status", "bcrm_pn", "Choice", "bcrm_eggdetail", "1:N per oocyte", "Yes", "new", "2PN normal; 0/1/3PN flagged"),
        ("Fertilisation", "PB Status", "bcrm_pb", "Choice", "bcrm_eggdetail", "1:N per oocyte", "No", "code", "Column exists on bcrm_eggdetail"),
        ("Fertilisation", "Insemination Type", "bcrm_inseminationtype", "Choice", "bcrm_eggdetail", "1:N per oocyte", "No", "code", ""),
        ("Fertilisation", "Fertilisation Time", "bcrm_fretilizationtime", "DateTime", "bcrm_eggdetail", "1:N", "No", "code", "Inherited spelling - keep on the wire"),
        ("Cycle roll-up", "Number Fertilised", "bcrm_numberfertilised", "Whole Number", "bcrm_fertilitytreatment", "roll-up onto cycle", "No", "code", ""),
      ])

sheet("8. Embryo Development",
      "Day 0-7 assessment. bcrm_egg (embryo) + bcrm_eggdetail (per-day rows).",
      [
        ("Embryo", "Embryo Seq No", "bcrm_seqno", "Whole Number", "bcrm_egg", "per-cycle sequence", "Yes", "new", ""),
        ("Embryo", "Source Oocyte", "bcrm_sourceoocyte", "Lookup", "bcrm_egg", "N:1 provenance", "No", "new", ""),
        ("Embryo", "Source Sperm", "bcrm_sourcesperm", "Lookup", "bcrm_egg / bcrm_semenprep", "N:1 provenance", "No", "new", ""),
        ("Assessment", "Assessment Day", "bcrm_day", "Choice", "bcrm_eggdetail", "1:N per embryo", "Yes", "code", "Day 2..7"),
        ("Assessment", "Cell Count", "bcrm_cellcount", "Whole Number", "bcrm_eggdetail", "field", "No", "code", ""),
        ("Assessment", "Expansion (1-6)", "bcrm_expansion", "Whole Number", "bcrm_eggdetail", "field", "No", "new", "Gardner"),
        ("Assessment", "ICM (A/B/C)", "bcrm_icm", "Choice", "bcrm_eggdetail", "field", "No", "code", "Gardner"),
        ("Assessment", "TE (A/B/C)", "bcrm_te", "Choice", "bcrm_eggdetail", "field", "No", "code", "Gardner"),
        ("Assessment", "Fragmentation %", "bcrm_fragmentation", "Whole Number", "bcrm_eggdetail", "field", "No", "code", ""),
        ("Assessment", "Development Stage", "bcrm_blastocyst / bcrm_morula", "Choice", "bcrm_eggdetail", "field", "No", "code", "Istanbul consensus stages"),
        ("Assessment", "Fate", "bcrm_embryofate", "Choice", "bcrm_eggdetail", "field", "Yes", "code", "Continue/Freeze/Transfer/Biopsy/Discard/Donate"),
        ("Assessment", "Symmetry", "bcrm_symmetry", "Choice", "bcrm_eggdetail", "field", "No", "code", ""),
        ("Assessment", "Comment", "bcrm_comment", "Multiline Text", "bcrm_eggdetail", "field", "No", "code", ""),
      ])

sheet("9. Embryo Transfer",
      "Fresh/FET transfer. New entity bcrm_embryotransfer (01-DATA-MODEL.md 3.7).",
      [
        ("Transfer", "Transfer ID", "bcrm_embryotransferid", "AutoNumber", "bcrm_embryotransfer", "primary key", "Yes", "new", "ET-{seq}"),
        ("Transfer", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1", "Yes", "new", ""),
        ("Transfer", "Transfer Date", "bcrm_transferdate", "DateTime", "bcrm_embryotransfer", "field", "Yes", "new", ""),
        ("Transfer", "Embryo", "bcrm_embryo", "Lookup", "bcrm_egg", "N:1 -> embryo", "Yes", "new", ""),
        ("Transfer", "Transfer Type", "bcrm_transfertype", "Choice", "bcrm_embryotransfer", "field", "Yes", "new", "Fresh / Frozen (FET)"),
        ("Transfer", "Transfer Order", "bcrm_transferorder", "Whole Number", "bcrm_embryotransfer", "field", "No", "new", ""),
        ("Transfer", "Catheter", "bcrm_catheter", "Text", "bcrm_embryotransfer", "field", "No", "new", ""),
        ("People", "Embryologist", "bcrm_embryologist", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("People", "Doctor", "bcrm_doctor", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("People", "Witness", "bcrm_witness", "Lookup", "systemuser / bcrm_staff", "N:1 (independent)", "Yes", "new", ""),
        ("Transfer", "Outcome", "bcrm_outcome", "Choice", "bcrm_embryotransfer", "field", "No", "new", ""),
        ("Cycle field", "No of Embryos to be Replaced", "bcrm_no_of_embryos_to_be_replaced", "Whole Number", "bcrm_fertilitytreatment", "field on cycle", "No", "code", ""),
      ])

sheet("10. IUI - DI",
      "Intrauterine insemination. Reuses bcrm_semenprep + bcrm_insemination (method = IUI).",
      [
        ("IUI", "Insemination ID", "bcrm_inseminationid", "AutoNumber", "bcrm_insemination", "primary key", "Yes", "new", ""),
        ("IUI", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1", "Yes", "new", ""),
        ("IUI", "Method", "bcrm_method", "Choice", "bcrm_insemination", "field", "Yes", "new", "IUI / DI"),
        ("IUI", "Date", "bcrm_date", "DateTime", "bcrm_insemination", "field", "Yes", "new", ""),
        ("IUI", "Prepared Sperm (source)", "bcrm_semenprep", "Lookup", "bcrm_semenprep", "N:1", "Yes", "new", ""),
        ("IUI", "Catheter", "bcrm_catheter", "Text", "bcrm_insemination", "field", "No", "new", ""),
        ("People", "Clinician", "bcrm_embryologist / bcrm_doctor", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("IUI", "Outcome", "bcrm_outcome", "Choice", "bcrm_insemination", "field", "No", "new", ""),
      ])

sheet("11. Gametes",
      "Specimen master view (egg/sperm/embryo). Primary entity: bcrm_egg (filtered by type).",
      [
        ("Specimen", "Gamete/Specimen ID", "bcrm_eggs_id", "AutoNumber", "bcrm_egg", "primary key", "Yes", "code", ""),
        ("Specimen", "Sequence No", "bcrm_seqno", "Whole Number", "bcrm_egg", "per-cycle sequence", "No", "new", ""),
        ("Specimen", "Type", "bcrm_type", "Choice", "bcrm_egg", "field", "Yes", "code", "Egg / Sperm / Embryo"),
        ("Specimen", "Patient", "bcrm_patient", "Lookup", "contact", "N:1", "Yes", "code", ""),
        ("Specimen", "Partner", "bcrm_partner", "Lookup", "contact", "N:1", "No", "code", ""),
        ("Specimen", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "N:1", "Yes", "code", ""),
        ("Status", "Current Status", "bcrm_egg_status", "Choice", "bcrm_egg", "field", "No", "code", "Collected/Frozen/Thawed/Used/Discarded"),
        ("Storage", "Storage Location", "bcrm_location", "Text / Lookup", "bcrm_egg / bcrm_cryolocation", "N:1 -> location", "No", "code", ""),
        ("Storage", "Freeze Date", "bcrm_freeze_date", "DateTime", "bcrm_egg", "field", "No", "code", ""),
        ("Storage", "Thaw Date", "bcrm_datetimeofeggthaw", "DateTime", "bcrm_egg", "field", "No", "code", ""),
        ("Storage", "Tank / Goblet / Straw", "bcrm_nitrogen_tank / bcrm_goblet / bcrm_straw", "Text", "bcrm_egg", "field", "No", "code", "Or via bcrm_cryolocation"),
        ("Consent", "Consent / Expiry", "bcrm_consent / bcrm_consent_expires_on", "Text / DateTime", "bcrm_egg", "field", "No", "code", ""),
      ])

sheet("12. Cryo Storage",
      "Tank graphical view + position hierarchy + inventory. bcrm_cryotank + bcrm_cryolocation (new); inventory derived over bcrm_egg.",
      [
        ("Tank", "Tank ID", "bcrm_cryotankid", "AutoNumber", "bcrm_cryotank", "primary key", "Yes", "new", "TANK-{seq}"),
        ("Tank", "Tank Name", "bcrm_name", "Text", "bcrm_cryotank", "field", "Yes", "new", ""),
        ("Tank", "Site", "bcrm_site", "Lookup", "bcrm_site / account", "N:1 -> site", "No", "new", ""),
        ("Tank", "Capacity", "bcrm_capacity", "Whole Number", "bcrm_cryotank", "field", "No", "new", ""),
        ("Tank", "Current Usage", "bcrm_currentusage", "Whole Number", "bcrm_cryotank", "roll-up / maintained", "No", "new", "Drives the fill % gauge"),
        ("Tank", "Active", "bcrm_active", "Two Options", "bcrm_cryotank", "field", "No", "new", ""),
        ("Tank", "Status", "bcrm_status", "Choice", "bcrm_cryotank", "field", "No", "new", "OK / Low N2 / Maintenance / Alarm"),
        ("Location", "Location ID", "bcrm_cryolocationid", "AutoNumber", "bcrm_cryolocation", "primary key", "Yes", "new", "LOC-{seq}"),
        ("Location", "Tank", "bcrm_tank", "Lookup", "bcrm_cryotank", "N:1 -> tank", "Yes", "new", ""),
        ("Location", "Canister Number", "bcrm_canisternumber", "Text / Whole Number", "bcrm_cryolocation", "field", "No", "new", ""),
        ("Location", "Goblet Number", "bcrm_gobletnumber", "Text / Whole Number", "bcrm_cryolocation", "field", "No", "new", ""),
        ("Location", "Straw/Vial Number", "bcrm_strawvialnumber", "Text / Whole Number", "bcrm_cryolocation", "field", "No", "new", ""),
        ("Location", "Storage Type", "bcrm_storagetype", "Choice", "bcrm_cryolocation", "field", "No", "new", "Embryo / Sperm / Oocyte"),
        ("Location", "Storage Status", "bcrm_storagestatus", "Choice", "bcrm_cryolocation", "field", "No", "new", "Occupied / Available / Reserved"),
        ("Inventory", "Specimen (occupant)", "bcrm_egg", "Lookup", "bcrm_egg", "1:N frozen specimens", "No", "new", "Inventory = query over frozen bcrm_egg by tank"),
        ("Inventory", "Patient / Cycle", "bcrm_patient / bcrm_treatment_cycle", "Lookup", "contact / cycle", "via bcrm_egg", "No", "code", ""),
      ])

sheet("13. Freeze - Thaw",
      "Cryopreservation & warming events. bcrm_freezeevent / bcrm_thawevent (new).",
      [
        ("Freeze", "Freeze Event ID", "bcrm_freezeeventid", "AutoNumber", "bcrm_freezeevent", "primary key", "Yes", "new", "FRZ-{seq}"),
        ("Freeze", "Material Record", "bcrm_materialrecord", "Lookup", "bcrm_egg", "N:1 -> specimen", "Yes", "new", ""),
        ("Freeze", "Material Type", "bcrm_materialtype", "Choice", "bcrm_freezeevent", "field", "Yes", "new", "Embryo/Sperm/Oocyte"),
        ("Freeze", "Freeze Date", "bcrm_freezedate", "DateTime", "bcrm_freezeevent", "field", "Yes", "new", ""),
        ("Freeze", "Storage Location", "bcrm_storagelocation", "Lookup", "bcrm_cryolocation", "N:1 -> location", "Yes", "new", "Sets location Occupied"),
        ("Freeze", "Embryologist / Witness", "bcrm_embryologist / bcrm_witness", "Lookup", "systemuser / bcrm_staff", "N:1", "Yes", "new", ""),
        ("Thaw", "Thaw Event ID", "bcrm_thaweventid", "AutoNumber", "bcrm_thawevent", "primary key", "Yes", "new", "THW-{seq}"),
        ("Thaw", "Material Record", "bcrm_materialrecord", "Lookup", "bcrm_egg", "N:1", "Yes", "new", ""),
        ("Thaw", "Scheduled / Actual Thaw Date", "bcrm_scheduledthawdate / bcrm_actualthawdate", "DateTime", "bcrm_thawevent", "field", "No", "new", ""),
        ("Thaw", "Quantity Thawed / Survived", "bcrm_quantitythawed / bcrm_quantitysurvived", "Whole Number", "bcrm_thawevent", "field", "No", "new", "Survived <= Thawed"),
        ("Thaw", "Outcome", "bcrm_outcome", "Choice", "bcrm_thawevent", "field", "No", "new", "Successful / Partial / Failed"),
      ])

sheet("14. Outcome Data",
      "Pregnancy test + delivery outcome. bcrm_pregnancy_outcome_details + cycle outcome fields.",
      [
        ("Early Outcome", "Date/Time of Early Outcome", "bcrm_date_time_of_early_outcome", "DateTime", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Early Outcome", "Result of Pregnancy Test", "bcrm_result_of_pregnancy_test", "Choice", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Early Outcome", "Pregnant", "bcrm_pregnant", "Two Options", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Early Outcome", "Outcome", "bcrm_outcome", "Choice / Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Delivery", "Delivery Date", "bcrm_delivery_date", "DateTime", "bcrm_pregnancy_outcome_details", "1:N -> cycle & patient", "No", "code", ""),
        ("Delivery", "Gestation", "bcrm_gestation", "Whole Number", "bcrm_pregnancy_outcome_details", "field", "No", "code", ""),
        ("Delivery", "Gender", "bcrm_gender", "Choice", "bcrm_pregnancy_outcome_details", "field", "No", "code", ""),
        ("Delivery", "Weight (grams)", "bcrm_weight_in_grams", "Whole Number", "bcrm_pregnancy_outcome_details", "field", "No", "code", ""),
        ("Delivery", "Forenames / Surname", "bcrm_forenames / bcrm_surname", "Text", "bcrm_pregnancy_outcome_details", "field", "No", "code", ""),
        ("Delivery", "Town / Country of Birth", "bcrm_town_or_district_of_birth / bcrm_country_of_birth", "Text", "bcrm_pregnancy_outcome_details", "field", "No", "code", ""),
        ("Cycle roll-up", "Embryos Transferred / Frozen", "bcrm_numberofembryostransferred / bcrm_numberofembryosfrozendayoffreeze", "Whole Number", "bcrm_fertilitytreatment", "field", "No", "code", ""),
      ])

sheet("15. Checklists - Consent",
      "Prerequisite gate. New entity bcrm_consentchecklist (01-DATA-MODEL.md 3.12).",
      [
        ("Checklist", "Checklist ID", "bcrm_consentchecklistid", "AutoNumber", "bcrm_consentchecklist", "primary key", "Yes", "new", "CCK-{seq}"),
        ("Checklist", "Treatment Cycle", "bcrm_treatment_cycle", "Lookup", "bcrm_treatment_cycle", "1:1 logical", "Yes", "new", ""),
        ("Checklist", "Consent Complete", "bcrm_consentcomplete", "Two Options", "bcrm_consentchecklist", "field", "Yes", "new", ""),
        ("Checklist", "Screening Complete", "bcrm_screeningcomplete", "Two Options", "bcrm_consentchecklist", "field", "Yes", "new", ""),
        ("Checklist", "Doctor Approval", "bcrm_doctorapproval", "Two Options", "bcrm_consentchecklist", "field", "Yes", "new", ""),
        ("Checklist", "Ready For Procedure", "bcrm_readyforprocedure", "Two Options", "bcrm_consentchecklist", "derived AND of the three", "No", "new", "Gates OPU/insemination/transfer"),
        ("Checklist", "Checked Date", "bcrm_checkeddate", "DateTime", "bcrm_consentchecklist", "field", "No", "new", ""),
        ("Checklist", "Checked By", "bcrm_checkedby", "Lookup", "systemuser / bcrm_staff", "N:1", "No", "new", ""),
      ])

sheet("16. Recommendation",
      "Clinical recommendations (largely bcrm_fertilitytreatment text fields).",
      [
        ("Recommendation", "ET Advice", "bcrm_et_advice", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Recommendation", "Scratch ET Details", "bcrm_scratch_et_details", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Recommendation", "Date of First Scratch", "bcrm_date_of_first_scratch", "DateTime", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Recommendation", "Treatment Features", "bcrm_treatmentfeatures", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("Recommendation", "Status Patient", "bcrm_status_patient", "Choice / Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
      ])

sheet("17. General Information",
      "Misc administrative / historical.",
      [
        ("General", "Centre", "bcrm_centre", "Text / Lookup", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("General", "Referring Centre", "bcrm_referring_centre", "Text / Lookup", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("General", "Any Cycle Complications", "bcrm_anycyclecomplications", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("General", "Year of Treatment", "bcrm_yearoftreatment", "Text / Whole Number", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("General", "Description", "bcrm_description", "Multiline Text", "bcrm_fertilitytreatment", "field", "No", "code", ""),
        ("General", "PGD/PGT", "bcrm_pgdpgd", "Two Options", "bcrm_fertilitytreatment", "field", "No", "code", ""),
      ])

# ---------------- Entities & Relationships sheet ----------------
er = wb.create_sheet("Entities & Relationships")
er.sheet_view.showGridLines = False
er["A1"] = "Entities & relationships"; er["A1"].font = title_font
er["A2"] = "How the tabs' source entities relate to the Treatment Cycle. (new) = created by the Embryo Module."
er["A2"].font = sub_font
ecols = ["Entity (logical)", "Represents", "Relationship to bcrm_treatment_cycle", "New?"]
ewid = [34, 34, 46, 8]
for j, c in enumerate(ecols, 1):
    cell = er.cell(4, j, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
    er.column_dimensions[get_column_letter(j)].width = ewid[j-1]
ents = [
    ("bcrm_treatment_cycle", "The treatment cycle (this form)", "-", ""),
    ("bcrm_fertilitytreatment", "Stimulation / folliculogram / outcome detail", "1:1 (or N:1) with cycle", ""),
    ("bcrm_pregnancy_outcome_details", "Delivery / baby outcome", "1:N -> cycle & patient", ""),
    ("contact", "Patient / Partner / Donor", "N:1 (bcrm_patient, bcrm_partner)", ""),
    ("account / bcrm_clinic", "Managing / Procedure clinic", "N:1", ""),
    ("systemuser / bcrm_staff", "Doctor / Nurse / Embryologist / Witness", "N:1 (several role lookups)", ""),
    ("bcrm_eventtemplate + detail", "Event/display template per treatment type", "N:1; ApplyToCycle seeds events", "new"),
    ("bcrm_embryoevent", "Worklist / diary event", "1:N per cycle", "new"),
    ("bcrm_threepointcheck", "Identity/consent gate", "1:N per cycle / per event", "new"),
    ("bcrm_consentchecklist", "Prerequisite gate", "1:1 with cycle", "new"),
    ("bcrm_opu", "Egg collection", "1:N per cycle", "new"),
    ("bcrm_semenprep", "Semen preparation", "1:N per cycle", "new"),
    ("bcrm_insemination", "IVF/ICSI/IUI insemination", "1:N per cycle", "new"),
    ("bcrm_egg", "Specimen master (egg/sperm/embryo)", "1:N per cycle", ""),
    ("bcrm_eggdetail", "Per-day assessment (denude/fert/dev)", "1:N per specimen", ""),
    ("bcrm_embryotransfer", "Transfer procedure", "1:N per cycle", "new"),
    ("bcrm_freezeevent / bcrm_thawevent", "Freeze / warming events", "1:N per cycle; N:1 -> specimen", "new"),
    ("bcrm_biopsy", "PGT biopsy / re-biopsy", "N:1 -> embryo (bcrm_egg)", "new"),
    ("bcrm_cryotank", "Storage tank", "referenced by frozen specimens", "new"),
    ("bcrm_cryolocation", "Tank position (canister/goblet/straw)", "N:1 -> tank; referenced by bcrm_egg", "new"),
    ("bcrm_researchproject", "Research project", "N:1 from cycle (banner/enrolment)", "new"),
    ("AppointmentDosage", "Drug schedule / doses", "1:N per cycle", ""),
]
for i, r in enumerate(ents):
    rr = 5 + i
    for j, v in enumerate(r, 1):
        cell = er.cell(rr, j, v); cell.alignment = wrap; cell.border = border
        if rr % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ZEBRA)
        if j == 4 and v == "new":
            cell.font = Font(color=DEEP, bold=True)
er.freeze_panes = "A5"

# Legend on the index sheet
lr = 5 + len(TABS) + 2
ov.cell(lr, 1, "Src legend:").font = Font(bold=True)
ov.cell(lr + 1, 1, "code = logical name confirmed in OxArBackendReact code")
ov.cell(lr + 2, 1, "form = label from the Treatment Cycle form (confirm logical name in Dataverse)")
ov.cell(lr + 3, 1, "new  = new entity/field from the Embryo Module (01-DATA-MODEL.md)")

wb.save("/workspace/oxarv1/docs/OXar_Treatment_Cycle_Tab_Attributes.xlsx")
print("saved; sheets:", wb.sheetnames)
